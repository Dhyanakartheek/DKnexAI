"""Excel and PDF report generation."""
import io
import logging
from datetime import datetime, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from db.models import Lead, LeadStatus, EmailType

logger = logging.getLogger(__name__)


async def generate_excel_report(db: AsyncSession) -> bytes:
    """Generate a weekly leads report as an Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    result = await db.execute(select(Lead).order_by(Lead.created_at.desc()))
    leads = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Tracker"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Lead ID", "Type", "Client", "Project / Role", "Skills",
        "Experience", "Deadline", "Status", "Assigned To",
        "Contact", "Contact Email", "Notes", "Created At"
    ]
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F1F5F9")
    for i, lead in enumerate(leads, 2):
        row = [
            lead.lead_id,
            lead.type,
            lead.client_name or "",
            lead.project_name or lead.role or "",
            lead.skills or "",
            lead.experience or "",
            lead.submission_deadline.strftime("%Y-%m-%d") if lead.submission_deadline else "",
            lead.status,
            lead.assigned_to or "",
            lead.contact_person or "",
            lead.contact_email or "",
            lead.notes or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        ws.append(row)
        if i % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=i, column=col_num).fill = alt_fill
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=i, column=col_num).border = border

    # Auto-width
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Count"])
    ws2.append(["Total Leads", len(leads)])
    for t in EmailType:
        cnt = sum(1 for l in leads if l.type == t.value)
        ws2.append([f"  {t.value}", cnt])
    ws2.append([])
    for s in LeadStatus:
        cnt = sum(1 for l in leads if l.status == s.value)
        ws2.append([f"Status: {s.value}", cnt])
    ws2.append(["Report Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_pdf_report(db: AsyncSession) -> bytes:
    """Generate a weekly summary PDF report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    result = await db.execute(select(Lead).order_by(Lead.created_at.desc()))
    leads = result.scalars().all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()

    accent = colors.HexColor("#6366F1")
    dark = colors.HexColor("#1E293B")
    light = colors.HexColor("#F1F5F9")

    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=dark, fontSize=22, spaceAfter=4
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        textColor=colors.HexColor("#64748B"), fontSize=10, spaceAfter=16
    )

    elements = []
    elements.append(Paragraph("📧 Email Automation Agent", title_style))
    elements.append(Paragraph(
        f"Weekly Lead Report · Generated {datetime.now(timezone.utc).strftime('%B %d, %Y')}",
        sub_style
    ))
    elements.append(Spacer(1, 0.3*cm))

    # Summary table
    summary_data = [["Metric", "Value"]]
    summary_data.append(["Total Leads", str(len(leads))])
    for t in EmailType:
        cnt = sum(1 for l in leads if l.type == t.value)
        summary_data.append([t.value, str(cnt)])
    summary_data.append(["", ""])
    for s in LeadStatus:
        cnt = sum(1 for l in leads if l.status == s.value)
        summary_data.append([f"Status: {s.value}", str(cnt)])

    t_summary = Table(summary_data, colWidths=[10*cm, 4*cm])
    t_summary.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), accent),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t_summary)
    elements.append(Spacer(1, 0.8*cm))

    # Lead detail table (first 20)
    elements.append(Paragraph("Lead Details (Latest 20)", styles["Heading2"]))
    lead_data = [["Lead ID", "Type", "Client", "Status", "Deadline"]]
    for lead in leads[:20]:
        lead_data.append([
            lead.lead_id,
            lead.type,
            (lead.client_name or "")[:30],
            lead.status,
            lead.submission_deadline.strftime("%Y-%m-%d") if lead.submission_deadline else "–",
        ])

    t_leads = Table(lead_data, colWidths=[3.5*cm, 2*cm, 6*cm, 3*cm, 3*cm])
    t_leads.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), dark),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [light, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(t_leads)

    doc.build(elements)
    return buf.getvalue()
